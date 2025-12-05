import json
import csv
import uuid
import openpyxl
from collections import defaultdict
from openpyxl.styles import Font
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, TemplateView, View
from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.forms import AuthenticationForm
from .models import Item, Purchase, Category, SaleRecord, Profile
from .forms import SignUpForm, ItemForm, PurchaseForm, CategoryForm, UserProfileForm
from django.urls import reverse_lazy
from django.db.models import Sum, F, Count, Max, Q
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.contrib import messages
from django.contrib.auth import login
from django.core.serializers.json import DjangoJSONEncoder

class CustomLoginView(LoginView):
    template_name = 'inventory/login.html'
    redirect_authenticated_user = True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['signup_form'] = SignUpForm()
        context['active_panel'] = 'login' 
        return context

class SignUpView(CreateView):
    model = Item 
    form_class = SignUpForm
    template_name = 'inventory/login.html'
    success_url = reverse_lazy('dashboard')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['login_form'] = AuthenticationForm()
        context['active_panel'] = 'signup'
        return context

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return redirect(self.success_url)

    def form_invalid(self, form):
        return render(self.request, self.template_name, {
            'form': form,
            'login_form': AuthenticationForm(),
            'active_panel': 'signup'
        })

class HomeView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = timezone.now().date()
        
        all_items = Item.objects.filter(user=user)
        low_stock_items = all_items.filter(quantity__lt=10)
        low_stock_count = low_stock_items.count()
        total_inventory_value = sum(item.total_value for item in all_items)

        todays_sales_query = SaleRecord.objects.filter(user=user, date_sold__date=today)
        sales_today = todays_sales_query.aggregate(sum=Sum('total_price'))['sum'] or 0
        items_sold_today = todays_sales_query.aggregate(qty=Sum('quantity'))['qty'] or 0

        current_month_sales = SaleRecord.objects.filter(
            user=user, 
            date_sold__year=today.year, 
            date_sold__month=today.month
        )
        
        first_day_this_month = today.replace(day=1)
        last_day_prev_month = first_day_this_month - timedelta(days=1)
        
        previous_month_sales = SaleRecord.objects.filter(
            user=user,
            date_sold__year=last_day_prev_month.year,
            date_sold__month=last_day_prev_month.month
        )

        monthly_revenue = current_month_sales.aggregate(sum=Sum('total_price'))['sum'] or 0
        monthly_items_sold = current_month_sales.aggregate(qty=Sum('quantity'))['qty'] or 0
        
        monthly_profit = 0
        for sale in current_month_sales:
            cost_price = getattr(sale, 'unit_cost_at_sale', 0)
            if cost_price == 0:
                cost_price = sale.product.average_cost
            total_cost = cost_price * sale.quantity
            monthly_profit += (sale.total_price - total_cost)

        current_month_records = current_month_sales.select_related('product').order_by('-date_sold')
        previous_month_records = previous_month_sales.select_related('product').order_by('-date_sold')

        dates_30 = []
        sales_30 = []
        for i in range(30):
            d = today - timedelta(days=29-i)
            day_sales = SaleRecord.objects.filter(user=user, date_sold__date=d).aggregate(total=Sum('total_price'))['total'] or 0
            dates_30.append(d.strftime('%b %d'))
            sales_30.append(int(day_sales))

        dates_7 = dates_30[-7:]
        sales_7 = sales_30[-7:]

        top_products_query = current_month_sales.values('product__name').annotate(total_qty=Sum('quantity')).order_by('-total_qty')[:5]
        top_products_labels = [item['product__name'] for item in top_products_query]
        top_products_data = [item['total_qty'] for item in top_products_query]

        category_query = current_month_sales.values('product__category__name').annotate(total=Sum('total_price')).order_by('-total')
        category_labels = [item['product__category__name'] for item in category_query]
        category_data = [item['total'] for item in category_query]

        context.update({
            'low_stock_count': low_stock_count,
            'low_stock_items': low_stock_items,
            'total_inventory_value': f"{int(total_inventory_value):,}",
            'sales_today': f"{int(sales_today):,}",
            'items_sold_today': items_sold_today,
            'monthly_revenue': f"{int(monthly_revenue):,}",
            'monthly_items_sold': monthly_items_sold,
            'monthly_profit': f"{int(monthly_profit):,}",
            'graph_dates_30': json.dumps(dates_30),
            'graph_sales_30': json.dumps(sales_30),
            'graph_dates_7': json.dumps(dates_7),
            'graph_sales_7': json.dumps(sales_7),
            'top_products_labels': json.dumps(top_products_labels),
            'top_products_data': json.dumps(top_products_data),
            'category_labels': json.dumps(category_labels),
            'category_data': json.dumps(category_data),
            'current_month_records': current_month_records,
            'previous_month_records': previous_month_records,
            'current_month_name': today.strftime('%B'),
            'prev_month_name': last_day_prev_month.strftime('%B'),
        })
        return context

class ProfileView(LoginRequiredMixin, View):
    template_name = 'inventory/profile.html'

    def get(self, request):
        if not hasattr(request.user, 'profile'):
            Profile.objects.create(user=request.user)

        profile_form = UserProfileForm(instance=request.user, user=request.user)
        return render(request, self.template_name, {
            'profile_form': profile_form
        })

    def post(self, request):
        if not hasattr(request.user, 'profile'):
            Profile.objects.create(user=request.user)

        profile_form = UserProfileForm(request.POST, instance=request.user, user=request.user)
        if profile_form.is_valid():
            profile_form.save()
            messages.success(request, "Profile updated successfully!")
            return redirect('profile')
        
        return render(request, self.template_name, {
            'profile_form': profile_form
        })

class ProductListView(LoginRequiredMixin, ListView):
    model = Item
    template_name = 'inventory/product_list.html'
    context_object_name = 'items'

    def get_queryset(self):
        query = self.request.GET.get('q')
        qs = Item.objects.filter(user=self.request.user)
        if query:
            qs = qs.filter(name__icontains=query)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(user=self.request.user)
        return context

class SalesBookView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/sales_book.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('q')
        
        sales_qs = SaleRecord.objects.filter(user=self.request.user).select_related('product', 'product__category').order_by('-date_sold')
        
        if query:
            sales_qs = sales_qs.filter(
                Q(order_id__icontains=query) | 
                Q(product__name__icontains=query)
            )

        grouped_orders = defaultdict(lambda: {
            'items': [], 
            'total_amount': 0, 
            'total_qty': 0, 
            'date': None,
            'order_id': None,
            'is_legacy': False
        })

        sorted_order_ids = []

        for sale in sales_qs:
            if not sale.order_id:
                oid = f"LEGACY-{sale.pk}"
                is_legacy = True
            else:
                oid = sale.order_id
                is_legacy = False
            
            if oid not in grouped_orders:
                grouped_orders[oid]['order_id'] = sale.order_id if not is_legacy else "N/A"
                grouped_orders[oid]['date'] = sale.date_sold
                grouped_orders[oid]['is_legacy'] = is_legacy
                sorted_order_ids.append(oid)
            
            grouped_orders[oid]['items'].append(sale)
            grouped_orders[oid]['total_amount'] += sale.total_price
            grouped_orders[oid]['total_qty'] += sale.quantity

        receipts_list = []
        for oid in sorted_order_ids:
            receipts_list.append(grouped_orders[oid])

        context['receipts'] = receipts_list
        context['search_query'] = query
        return context

class AddProductView(LoginRequiredMixin, CreateView):
    model = Item
    form_class = ItemForm
    template_name = 'inventory/add_item.html'
    success_url = reverse_lazy('product_list')
    
    def form_valid(self, form):
        try:
            if form.cleaned_data.get('quantity') is not None and form.cleaned_data.get('quantity') < 0:
                form.add_error('quantity', "Quantity must be positive")
                return self.form_invalid(form)
            if form.cleaned_data.get('selling_price') is not None and form.cleaned_data.get('selling_price') < 0:
                form.add_error('selling_price', "Price must be positive")
                return self.form_invalid(form)
        except (ValueError, TypeError):
            return self.form_invalid(form)

        form.instance.user = self.request.user
        self.object = form.save()
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Product added successfully!'})
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
        return super().form_invalid(form)

class AddPurchaseView(LoginRequiredMixin, CreateView):
    model = Purchase
    form_class = PurchaseForm
    template_name = 'inventory/add_purchase.html'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

class SaleView(LoginRequiredMixin, View):
    template_name = 'inventory/sale.html'

    def get(self, request):
        recent_sales = SaleRecord.objects.filter(user=request.user, date_sold__date=timezone.now().date()).order_by('-date_sold')
        items = Item.objects.filter(user=request.user).values('id', 'name', 'selling_price', 'quantity')
        products_json = json.dumps(list(items), cls=DjangoJSONEncoder)
        return render(request, self.template_name, {
            'recent_sales': recent_sales,
            'products_json': products_json
        })

    def post(self, request):
        try:
            data = json.loads(request.body)
            cart_items = data.get('items', [])
            if not cart_items:
                return JsonResponse({'status': 'error', 'message': 'Cart is empty'}, status=400)

            order_id = f"ORD-{uuid.uuid4().hex[:8].upper()}"

            with transaction.atomic():
                for item_data in cart_items:
                    product_id = item_data.get('id')
                    qty = int(item_data.get('qty'))
                    
                    product = Item.objects.select_for_update().get(pk=product_id, user=request.user)
                    
                    if product.quantity < qty:
                        raise ValueError(f"Insufficient stock for {product.name}")
                    
                    total_price = product.selling_price * qty
                    
                    sale_data = {
                        'product': product,
                        'quantity': qty,
                        'total_price': total_price,
                        'user': request.user,
                        'order_id': order_id 
                    }
                    
                    if hasattr(SaleRecord, 'unit_cost_at_sale'):
                        sale_data['unit_cost_at_sale'] = product.average_cost

                    SaleRecord.objects.create(**sale_data)
                    
                    product.quantity -= qty
                    product.save()
            
            messages.success(request, f"Sale completed! Tracking #: {order_id}")
            return JsonResponse({'status': 'success', 'redirect': reverse_lazy('sales')})

        except Item.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': 'An error occurred processing the sale.'}, status=500)

def delete_sale(request, pk):
    with transaction.atomic():
        sale = get_object_or_404(SaleRecord, pk=pk, user=request.user)
        try:
            item = Item.objects.select_for_update().get(pk=sale.product.pk)
            item.quantity += sale.quantity
            item.save()
        except Item.DoesNotExist:
            pass
        
        sale.delete()
        messages.success(request, "Sale reversed and stock restored.")
    return redirect('sales')

def delete_item(request, pk):
    item = get_object_or_404(Item, pk=pk, user=request.user)
    item.delete()
    messages.success(request, "Item deleted successfully.")
    return redirect('product_list')

def export_daily_sales(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="daily_sales.csv"'
    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Date', 'Product Name', 'Qty Sold', 'Unit Price', 'Total Price'])
    sales = SaleRecord.objects.filter(user=request.user, date_sold__date=timezone.now().date())
    for sale in sales:
        writer.writerow([sale.order_id, sale.date_sold, sale.product.name, sale.quantity, sale.product.selling_price, sale.total_price])
    return response

def export_monthly_sales(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Monthly_Sales_Report.xlsx"'

    wb = openpyxl.Workbook()
    
    def create_report_sheet(workbook, title, sales_data):
        if workbook.active.title == "Sheet":
            ws = workbook.active
            ws.title = title
        else:
            ws = workbook.create_sheet(title=title)
        
        headers = ['Order ID', 'Product / Item', 'Category', 'Date of Sale', 'Month', 'Year', 'Qty Sold', 'Gross Sales', 'Total Cost', 'Net Profit']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for sale in sales_data:
            revenue = sale.total_price
            cost_price = getattr(sale, 'unit_cost_at_sale', 0)
            if cost_price == 0:
                cost_price = sale.product.average_cost
                
            total_cost = cost_price * sale.quantity
            profit = revenue - total_cost
            
            ws.append([
                sale.order_id if sale.order_id else "N/A",
                sale.product.name,
                sale.product.category.name,
                sale.date_sold.strftime('%Y-%m-%d'),
                sale.date_sold.strftime('%B'),
                sale.date_sold.year,
                sale.quantity,
                revenue,
                total_cost,
                profit
            ])

    today = timezone.now().date()
    first_day_this_month = today.replace(day=1)
    last_day_prev_month = first_day_this_month - timedelta(days=1)

    current_sales = SaleRecord.objects.filter(
        user=request.user, 
        date_sold__month=today.month, 
        date_sold__year=today.year
    ).select_related('product', 'product__category').order_by('-date_sold')
    
    create_report_sheet(wb, f"Current ({today.strftime('%B')})", current_sales)

    prev_sales = SaleRecord.objects.filter(
        user=request.user, 
        date_sold__month=last_day_prev_month.month, 
        date_sold__year=last_day_prev_month.year
    ).select_related('product', 'product__category').order_by('-date_sold')
    
    create_report_sheet(wb, f"Previous ({last_day_prev_month.strftime('%B')})", prev_sales)

    wb.save(response)
    return response

class AddCategoryView(LoginRequiredMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/add_category.html'
    success_url = reverse_lazy('product_list')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)